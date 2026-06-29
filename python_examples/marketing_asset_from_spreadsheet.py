"""
Wrapper example: drive marketing-asset and scenario creation from a spreadsheet.

For each row in the spreadsheet this script:
  1. looks up the configured AI workflow (by name) for its model
  2. finds the Model and Footwear records in the supporting_assets library
  3. creates a parent marketing asset, setting the customer-specific custom fields
  4. creates up to 3 scenarios under that asset, each combining:
       - detail images uploaded from the row's "Detail path" folder
       - the model's and footwear's view image (front/back/side), each tagged with the
         workflow image field ("wf_field") it maps to, so the scenario can be replayed
         or copied in the app
       - a prompt taken from the scenario's configured prompt template (by name)

The spreadsheet is expected to have a header row with these columns:

    fieldA       customer-specific field (changes per environment)
    fieldB       customer-specific field (changes per environment)
    Model        name of a 'base_model' record in the supporting_assets library
    Footwear     name of a 'supporting_product' record in the supporting_assets library
    Detail path  folder containing the detail images for the row

This example composes the smaller single-purpose examples:
    find_library_record.find_library_record_by_name
    find_prompt_template.find_prompt_template_by_name
    upload_file.upload_file
    create_marketing_asset.create_marketing_asset
    create_scenario.create_scenario
    get_workflow.get_workflow_by_name
"""
import glob
import os

from openpyxl import load_workbook

from create_marketing_asset import create_marketing_asset
from create_scenario import create_scenario
from find_library_record import find_library_record_by_name
from find_prompt_template import find_prompt_template_by_name
from get_workflow import get_workflow_by_name
from upload_file import upload_file

# ---------------------------------------------------------------------------
# Program configuration (adjust per environment / run)
# ---------------------------------------------------------------------------
# default to the sample spreadsheet bundled next to this script, so it is found regardless
# of the current working directory (override with the SPREADSHEET_PATH environment variable)
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SPREADSHEET_PATH = os.getenv("SPREADSHEET_PATH", os.path.join(SCRIPT_DIR, "marketing_assets.xlsx"))

# a relative "Detail path" in the spreadsheet is resolved against this directory, so the
# script works regardless of the current working directory. Defaults to the folder the
# spreadsheet lives in (override with DETAIL_BASE_DIR). Absolute Detail paths are used as-is.
DETAIL_BASE_DIR = os.getenv("DETAIL_BASE_DIR", os.path.dirname(os.path.abspath(SPREADSHEET_PATH)))

# the AI workflow whose model every scenario will use
WORKFLOW_NAME = "Marketing Asset Integration"

# the prompt template (by name) whose prompt_text is used for each scenario's prompt
SCENARIO_1_PROMPT_TEMPLATE = "On-Body Studio - Front"
SCENARIO_2_PROMPT_TEMPLATE = "On-Body Studio - Back"
SCENARIO_3_PROMPT_TEMPLATE = "On-Body Studio - Side"

# the library that holds the Model (base_model) and Footwear (supporting_product) records
LIBRARY_KEY = "supporting_assets"

# image size / count applied to every scenario
WIDTH = 1024
HEIGHT = 1024
NUM_IMAGES = 1

SCENARIO_TYPE = "3d_to_real"

# asset_type tag used in scenario input_file_ids. The uploaded detail images and the
# model/footwear view images all reference file records, so their entries are tagged as files.
ASSET_TYPE_FILE = "file"

# asset_type of the library records themselves (used to find them in the library)
RECORD_TYPE_MODEL = "base_model"
RECORD_TYPE_FOOTWEAR = "supporting_product"

# spreadsheet column headers
COL_FIELD_A = "season"
COL_FIELD_B = "style_number"
COL_MODEL = "Model"
COL_FOOTWEAR = "Footwear"
COL_DETAIL_PATH = "Detail path"

# the view image field to pull from the Model / Footwear records, per scenario
VIEW_FRONT = "front_view"
VIEW_BACK = "back_view"
VIEW_SIDE = "side_view"

# scenario definitions: detail-file suffixes, view field, required, and the prompt
# template (by name) supplying the scenario's prompt.
# scenario 1 is always created; 2 and 3 are skipped when their detail file is absent
SCENARIO_DEFINITIONS = [
    {"suffixes": [1, 2], "view": VIEW_FRONT, "required": True, "prompt_template": SCENARIO_1_PROMPT_TEMPLATE},
    {"suffixes": [4], "view": VIEW_BACK, "required": False, "prompt_template": SCENARIO_2_PROMPT_TEMPLATE},
    {"suffixes": [3], "view": VIEW_SIDE, "required": False, "prompt_template": SCENARIO_3_PROMPT_TEMPLATE},
]


def _file_id(value) -> str:
    """Library file fields may be returned as a joined dict or a plain id string."""
    if isinstance(value, dict):
        return value.get("_id")
    return value


def _detail_files(detail_path: str, suffix: int) -> list:
    """Return files in the folder whose name ends with the given numeric suffix."""
    return sorted(glob.glob(os.path.join(detail_path, f"*_{suffix}.*")))


def read_rows(spreadsheet_path: str) -> list:
    """Read the spreadsheet into a list of dicts keyed by the header row."""
    workbook = load_workbook(spreadsheet_path, read_only=True, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def build_input_file_ids(detail_path: str, suffixes: list, model_rec: dict,
                         footwear_rec: dict, view: str,
                         model_wf_field: str, footwear_wf_field: str) -> list:
    """Upload the row's detail images and combine them with the model/footwear views.

    The model and footwear entries carry a 'wf_field' set to a workflow image-field
    id, so the app can map them back to the workflow's inputs when the scenario is
    later replayed or copied. The detail images have no corresponding workflow field,
    so they are left unmapped.
    """
    input_file_ids = []

    # upload each detail image and capture its file id
    for suffix in suffixes:
        for file_path in _detail_files(detail_path, suffix):
            uploaded = upload_file(file_path)
            input_file_ids.append({"_id": uploaded["_id"], "asset_type": ASSET_TYPE_FILE})

    # pull the file id from the matching view field of the model and footwear records.
    # the view field resolves to a file record, so the entry is tagged as a file and
    # carries the workflow image field it maps to (order between the two does not matter)
    input_file_ids.append({"_id": _file_id(model_rec.get(view)),
                           "asset_type": ASSET_TYPE_FILE, "wf_field": model_wf_field})
    input_file_ids.append({"_id": _file_id(footwear_rec.get(view)),
                           "asset_type": ASSET_TYPE_FILE, "wf_field": footwear_wf_field})

    return input_file_ids


def process_row(row: dict, workflow: dict, model_id: str,
                model_wf_field: str, footwear_wf_field: str) -> None:
    # resolve a relative Detail path against DETAIL_BASE_DIR (absolute paths pass through)
    detail_path = os.path.join(DETAIL_BASE_DIR, row[COL_DETAIL_PATH])

    # find the model and footwear records in the supporting_assets library
    model_rec = find_library_record_by_name(LIBRARY_KEY, row[COL_MODEL], RECORD_TYPE_MODEL)
    if model_rec is None:
        raise ValueError(f"Model record not found: {row[COL_MODEL]!r}")

    footwear_rec = find_library_record_by_name(LIBRARY_KEY, row[COL_FOOTWEAR], RECORD_TYPE_FOOTWEAR)
    if footwear_rec is None:
        raise ValueError(f"Footwear record not found: {row[COL_FOOTWEAR]!r}")

    # create the parent marketing asset with the customer-specific custom fields
    asset = create_marketing_asset({
        COL_FIELD_A: row.get(COL_FIELD_A),
        COL_FIELD_B: row.get(COL_FIELD_B),
    })
    print(f"Created marketing asset {asset['_id']}")

    # create the scenarios
    for index, definition in enumerate(SCENARIO_DEFINITIONS, start=1):
        # optional scenarios are skipped when none of their detail files are present
        if not definition["required"]:
            has_files = any(_detail_files(detail_path, s) for s in definition["suffixes"])
            if not has_files:
                print(f"  Skipping scenario {index}: no detail file for suffix(es) "
                      f"{definition['suffixes']}")
                continue

        # resolve the scenario's prompt from its configured prompt template
        template = find_prompt_template_by_name(definition["prompt_template"])
        if template is None:
            raise ValueError(f"Prompt template not found: {definition['prompt_template']!r}")
        prompt = template.get("prompt_text", "")

        input_file_ids = build_input_file_ids(
            detail_path, definition["suffixes"], model_rec, footwear_rec, definition["view"],
            model_wf_field, footwear_wf_field)

        scenario = create_scenario(asset["_id"], {
            "prompt": prompt,
            "model": model_id,
            "workflow_id": workflow["_id"],
            "width": WIDTH,
            "height": HEIGHT,
            "num_images": NUM_IMAGES,
            "scenario_type": SCENARIO_TYPE,
            "input_file_ids": input_file_ids,
        })
        print(f"  Created scenario {index}: {scenario.get('_id')}")


def main() -> None:
    # look up the workflow once for the whole run
    workflow = get_workflow_by_name(WORKFLOW_NAME)
    if workflow is None:
        raise ValueError(f"Workflow not found: {WORKFLOW_NAME!r}")

    # use the workflow's AI model (first one if several are configured)
    models = workflow.get("models") or []
    if not models:
        raise ValueError(f"Workflow {WORKFLOW_NAME!r} has no model configured.")
    model_id = models[0]

    # the workflow's two image fields (e.g. "Supporting Asset 1" / "Supporting Asset 2")
    # are correlated to the model and footwear reference images. Their ids are read live
    # from the workflow, so nothing has to be hard-coded. Order does not matter.
    image_field_ids = [f.get("_id") for f in (workflow.get("image_fields") or []) if f.get("_id")]
    if len(image_field_ids) < 2:
        raise ValueError(f"Workflow {WORKFLOW_NAME!r} needs at least two image fields "
                         f"(found {len(image_field_ids)}).")
    model_wf_field, footwear_wf_field = image_field_ids[0], image_field_ids[1]

    rows = read_rows(SPREADSHEET_PATH)
    print(f"Read {len(rows)} row(s) from {SPREADSHEET_PATH}")

    for row in rows:
        process_row(row, workflow, model_id, model_wf_field, footwear_wf_field)


if __name__ == '__main__':
    main()
